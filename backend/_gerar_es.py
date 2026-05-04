import openpyxl, xlrd
import io, sys

TABUAS = {}

for ano, ext in [(2014,'xls'),(2015,'xls'),(2016,'xls'),(2017,'xls'),(2018,'xls'),(2019,'xls'),(2020,'xls'),(2021,'xlsx'),(2022,'xlsx'),(2023,'xlsx'),(2024,'xlsx')]:
    p = rf'C:\Users\Administrador\OneDrive\Documentos\Documents\previdenciario\backend\_ibge_{ano}.{ext}'
    try:
        valores = {}
        if ext == 'xls':
            wb = xlrd.open_workbook(p)
            sh = wb.sheet_by_index(0)
            for r in range(sh.nrows):
                idade_raw = sh.cell(r, 0).value
                if isinstance(idade_raw, (int, float)):
                    idade = int(idade_raw)
                    if 0 <= idade <= 100:
                        for c in range(sh.ncols - 1, -1, -1):
                            v = sh.cell(r, c).value
                            if isinstance(v, (int, float)) and 0 < v < 100:
                                valores[idade] = round(v, 2)
                                break
        else:
            wb = openpyxl.load_workbook(p, data_only=True)
            sh = wb.active
            for row in sh.iter_rows(values_only=True):
                if row and row[0] is not None:
                    idade_raw = row[0]
                    if isinstance(idade_raw, (int, float)):
                        idade = int(idade_raw)
                        if 0 <= idade <= 100:
                            for v in reversed(row):
                                if isinstance(v, (int, float)) and 0 < v < 100:
                                    valores[idade] = round(v, 2)
                                    break
        TABUAS[ano] = valores
    except Exception as e:
        print(f'ERRO {ano}: {e}')

DER_TO_TABUA = {
    2016: 2014, 2017: 2015, 2018: 2016, 2019: 2017,
    2020: 2018, 2021: 2019, 2022: 2020, 2023: 2021,
    2024: 2022, 2025: 2023, 2026: 2024,
}

linhas = []
for der_ano in sorted(DER_TO_TABUA.keys()):
    tabua_ano = DER_TO_TABUA[der_ano]
    if tabua_ano not in TABUAS:
        continue
    valores = TABUAS[tabua_ano]
    linhas.append(f'    {der_ano}: {{  # Tabua IBGE {tabua_ano} (publicada nov/{tabua_ano+1}, vigente DERs {der_ano})')
    for idade in range(50, 91):
        if idade in valores:
            linhas.append(f'        {idade}: _D("{valores[idade]:.2f}"),')
    linhas.append('    },')

with open(r'C:\Users\Administrador\OneDrive\Documentos\Documents\previdenciario\backend\_es_oficial_codigo.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(linhas))
print(f'Gerado _es_oficial_codigo.txt com {len(linhas)} linhas')
